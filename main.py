import openpyxl
import glob, os


def readAndSaveExcelFiles(source,destination,sheetNumber):

    wb1 = openpyxl.load_workbook(source)
    ws1 = wb1.worksheets[0]


    wb2 = openpyxl.load_workbook(destination)
    ws2 = wb2[wb2.sheetnames[-1]]

    mr = ws1.max_row
    mc = ws1.max_column
    informations = []
    data =[] 


    for i in range (2, mr + 1):
    
        data =[]
        for j in range (1, mc + 1):
            c = ws1.cell(row = i, column = j)
            data.append(c.value)

        informations.append(data)
        
    mrd =  ws2.max_row
    for i in informations:
        mrd = mrd+ 1 
        if(mrd>sheetNumber):
            wb2.create_sheet()
            ws2 = wb2[wb2.sheetnames[-1]]
            mrd = ws2.max_row
            
        
        ws2.append(i)
        


    wb2.save(str(destination))
    print(informations)


files = glob.glob("newExcels/*.xlsx")
files.sort(key=os.path.getmtime)


print("list all excel files : ")
print("\n".join(files))

mainExcel = "main.xlsx"


for i in files:
    readAndSaveExcelFiles(i,mainExcel,5)
    os.remove(i)





















































# from re import U
# import openpyxl as xl
# from openpyxl.xml.constants import MAX_ROW
# import glob, os


# def readAndSaveExcelFile(source,destantion,sheetNumber):
#     filename =source
#     wb1 = xl.load_workbook(filename)
#     ws1 = wb1.worksheets[0]
    

#     filename1 =destantion
#     wb2 = xl.load_workbook(filename1)
#     ws2 = wb2.active
    

#     mr = ws1.max_row
#     mc = ws1.max_column
#     u = []
#     uu =[] 


#     for i in range (2, mr + 1):
    
#         uu =[]
#         for j in range (1, mc + 1):
#             c = ws1.cell(row = i, column = j)
#             uu.append(c.value)
#             #ws2.cell(row = i, column = j).value = c.value
#         u.append(uu)
        
#     mrd = ws2.max_row
#     for data in u:
#         if(mrd==sheetNumber):
#             wb2.create_sheet()
#             ws2 = wb2[wb2.sheetnames[-1]]
#             mrd = ws2.max_row
#         ws2.append(data)
#         mrd = mrd+ 1    


#     wb2.save(str(filename1))
#     print(u)





# #os.chdir("/usr/bin/python3 /home/rain/Downloads/python/")


# files = glob.glob("newExcels/*.xlsx")
# files.sort(key=os.path.getmtime)
# print("list all excel files : ")
# print("\n".join(files))


# # for i in files:
# #     readAndSaveExcelFile(i,"2.xlsx",6)



# # for i in files:
# #     print(os.path.getatime(i))

# # for file in glob.glob("newExcels/*.xlsx"):
# #     print(file)
# #     print("created: %s" % time.ctime(os.path.getctime(file)))



